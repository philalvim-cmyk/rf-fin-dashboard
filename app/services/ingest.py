# app/services/ingest.py
import pandas as pd
import openpyxl
from utils.dates import to_datetime, competencia_mes

def stream_filter_by_competencia(file_path, sheet_name: str, competencia: str, columns: list[str]) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    missing = [c for c in columns if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Colunas ausentes na aba '{sheet_name}': {missing}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dt_emissao = to_datetime(row[idx["DATA_EMISSAO"]])
        comp = competencia_mes(dt_emissao)
        if comp != competencia:
            continue
        rows.append([row[idx[c]] for c in columns])

    wb.close()
    return pd.DataFrame(rows, columns=columns)