# app/services/train_history.py
from __future__ import annotations

import datetime as dt
from collections import defaultdict
from typing import Dict, Tuple, Optional

import openpyxl

from utils.strings import normalize_text
from db.repositories.history_repo import bulk_upsert_history


def train_history_from_excel(
    file_path: str,
    sheet_name: str,
    classificacao_col: str = "CLASSIFICAÇÃO RF",
    key_mode: str = "CONTA+PESSOA+CC",
    limit_rows: Optional[int] = None,
) -> dict:
    """
    Lê o Excel (read_only) e treina o histórico no SQLite.

    key_mode:
      - "CONTA+PESSOA+CC" (recomendado como base)
    """

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Aba '{sheet_name}' não encontrada. Abas: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Cabeçalho
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    required = ["NOME_CONTA", "NOME_PESSOA", "NOME_CENTRO_CUSTO", classificacao_col]
    missing = [c for c in required if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Colunas ausentes na aba '{sheet_name}': {missing}")

    # Contagem: (nome_conta, nome_pessoa, nome_cc, classificacao) -> hit_count
    counts: Dict[Tuple[str, str, str, str], int] = defaultdict(int)

    processed = 0
    skipped_empty_class = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        processed += 1
        if limit_rows and processed > limit_rows:
            break

        nome_conta = normalize_text(row[idx["NOME_CONTA"]])
        nome_pessoa = normalize_text(row[idx["NOME_PESSOA"]])
        nome_cc = normalize_text(row[idx["NOME_CENTRO_CUSTO"]])
        classificacao = normalize_text(row[idx[classificacao_col]])

        if not classificacao:
            skipped_empty_class += 1
            continue

        # chave completa
        key = (nome_conta, nome_pessoa, nome_cc, classificacao)
        counts[key] += 1

    wb.close()

    # Monta registros para bulk upsert
    now = dt.datetime.now().isoformat(timespec="seconds")
    records = []
    for (nome_conta, nome_pessoa, nome_cc, classificacao), hit_count in counts.items():
        records.append(
            {
                "key_type": key_mode,
                "nome_conta": nome_conta,
                "nome_pessoa": nome_pessoa,
                "nome_centro_custo": nome_cc,
                "classificacao_rf": classificacao,
                "hit_count": hit_count,
                "last_used_at": now,
            }
        )

    inserted_or_updated = bulk_upsert_history(records)

    return {
        "rows_processed": processed,
        "skipped_empty_class": skipped_empty_class,
        "unique_mappings": len(records),
        "db_rows_affected": inserted_or_updated,
    }