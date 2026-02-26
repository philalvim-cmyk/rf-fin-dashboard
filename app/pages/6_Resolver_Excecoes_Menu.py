import streamlit as st

st.set_page_config(page_title="Resolver Exceções (Menu)", layout="wide")

from app.utils.ui import render_sidebar_branding
render_sidebar_branding()

import pandas as pd
from pathlib import Path
import datetime as dt
import time

from app.utils.strings import normalize_text
from app.db.repositories.history_repo import bulk_upsert_history
from app.db.repositories.rubricas_repo import list_rubricas, ensure_rubricas_schema

from app.services.classify import apply_classification
from app.services.aggregate import consolidate_dinamica
from app.services.export import export_excel
from app.services.transform import normalize_keys, add_competencia

from openpyxl import load_workbook


def fmt_br(x) -> str:
    if x is None:
        return ""
    try:
        s = f"{float(x):,.2f}"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(x)


def apply_excel_number_format(path: Path) -> None:
    wb = load_workbook(path)
    num_fmt = "#,##0.00"

    if "DINAMICA_CONSOLIDADO" in wb.sheetnames:
        ws = wb["DINAMICA_CONSOLIDADO"]
        header = [c.value for c in ws[1]]
        if "VALOR" in header:
            col = header.index("VALOR") + 1
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_fmt

    if "DETALHADO_CLASSIFICADO" in wb.sheetnames:
        ws = wb["DETALHADO_CLASSIFICADO"]
        header = [c.value for c in ws[1]]
        if "VALOR_TITULO" in header:
            col = header.index("VALOR_TITULO") + 1
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_fmt

    wb.save(path)


def _progress(bar, status, pct: int, msg: str):
    pct = max(0, min(100, int(pct)))
    bar.progress(pct, text=f"{pct}% — {msg}")
    status.info(msg)


st.title("Resolver Exceções — selecionar rubrica no menu")

ensure_rubricas_schema()
rubricas = list_rubricas(only_active=True)
if not rubricas:
    st.warning("Nenhuma rubrica cadastrada. Vá em 'Plano de Contas — Rubricas' e cadastre primeiro.")
    st.stop()

with st.sidebar:
    st.header("Filtros de Rubricas")
    filtro_grupo = st.selectbox("Grupo", ["TODOS", "DESPESA", "RECEITA"])
    filtro_natureza = st.selectbox("Natureza", ["TODOS", "OPERACIONAL", "NAO_OPERACIONAL"])


def _filtrar_rubricas(lista):
    out = lista
    if filtro_grupo != "TODOS":
        out = [r for r in out if r.get("grupo") == filtro_grupo]
    if filtro_natureza != "TODOS":
        out = [r for r in out if r.get("natureza") == filtro_natureza]
    return out


rubricas_filtradas = _filtrar_rubricas(rubricas)
opcoes_display = [""] + [
    r["rubrica_display"].strip()
    for r in rubricas_filtradas
    if isinstance(r.get("rubrica_display"), str) and r["rubrica_display"].strip()
]
if len(opcoes_display) == 1:
    st.warning("Com esses filtros, não há rubricas. Ajuste os filtros no sidebar.")
    st.stop()

exports_dir = Path("data/exports")
exports = sorted(exports_dir.glob("export_rf_*.xlsx"))
if not exports:
    st.warning("Nenhum export encontrado em data/exports. Rode 'Aplicar Histórico' primeiro.")
    st.stop()

last_export = exports[-1]
st.caption(f"Usando último export: {last_export.name}")

df = pd.read_excel(last_export, sheet_name="DETALHADO_CLASSIFICADO", engine="openpyxl")

required_cols = ["NOME_CONTA", "NOME_PESSOA", "NOME_CENTRO_CUSTO", "VALOR_TITULO", "CLASSIFICACAO_RF"]
missing = [c for c in required_cols if c not in df.columns]
if missing:
    st.error(f"Export não tem as colunas esperadas: {missing}")
    st.stop()

for col_src, col_n in [
    ("NOME_CONTA", "NOME_CONTA_N"),
    ("NOME_PESSOA", "NOME_PESSOA_N"),
    ("NOME_CENTRO_CUSTO", "NOME_CC_N"),
]:
    if col_n not in df.columns:
        df[col_n] = df[col_src].fillna("").apply(normalize_text)

df_nc = df[df["CLASSIFICACAO_RF"] == "NAO_CLASSIFICADO"].copy()

st.write(f"Total linhas no export: {len(df)}")
st.write(f"Total NAO_CLASSIFICADO: {len(df_nc)}")

if df_nc.empty:
    st.success("Sem exceções!")
    st.stop()

grp = (
    df_nc.groupby(["NOME_CONTA_N", "NOME_PESSOA_N", "NOME_CC_N"], dropna=False)
        .agg(qtd=("VALOR_TITULO", "size"), valor=("VALOR_TITULO", "sum"))
        .reset_index()
        .sort_values(["valor", "qtd"], ascending=False)
)

grp["valor_br"] = grp["valor"].apply(fmt_br)

if "RUBRICA_ESCOLHIDA" not in grp.columns:
    grp["RUBRICA_ESCOLHIDA"] = ""

st.info("Selecione a rubrica no menu. Depois clique em 'Salvar lote'.")

view = grp[["NOME_CONTA_N", "NOME_PESSOA_N", "NOME_CC_N", "qtd", "valor_br", "RUBRICA_ESCOLHIDA"]].copy()

edited_view = st.data_editor(
    view,
    use_container_width=True,
    num_rows="dynamic",
    hide_index=True,
    column_config={
        "RUBRICA_ESCOLHIDA": st.column_config.SelectboxColumn(
            "Rubrica",
            options=opcoes_display,
            help="Selecione uma rubrica cadastrada no Plano de Contas."
        )
    }
)

col1, col2 = st.columns(2)

if "ultimo_export_reclassificado" not in st.session_state:
    st.session_state["ultimo_export_reclassificado"] = None

with col1:
    if st.button("Salvar lote no histórico (SQLite)"):
        now = dt.datetime.now().isoformat(timespec="seconds")
        to_save = edited_view[edited_view["RUBRICA_ESCOLHIDA"].astype(str).str.strip() != ""].copy()

        if to_save.empty:
            st.warning("Nada selecionado para salvar.")
            st.stop()

        records = []
        for _, r in to_save.iterrows():
            records.append({
                "key_type": "CONTA+PESSOA+CC",
                "nome_conta": r["NOME_CONTA_N"],
                "nome_pessoa": r["NOME_PESSOA_N"],
                "nome_centro_custo": r["NOME_CC_N"],
                "classificacao_rf": r["RUBRICA_ESCOLHIDA"].strip(),
                "hit_count": int(r["qtd"]),
                "last_used_at": now
            })

        affected = bulk_upsert_history(records)
        st.success(f"✅ Lote salvo! Registros gravados/atualizados no histórico: {affected}")
        st.info("Agora você pode reexportar aqui mesmo e baixar o Excel atualizado.")

    if st.button("Reexportar Excel com histórico atualizado (Download)"):
        bar = st.progress(0, text="0% — Iniciando reexportação...")
        status = st.empty()
        t0 = time.perf_counter()

        try:
            # ✅ CIRURGIA: removido show_time=True (incompatível)
            with st.spinner("Reclassificando e gerando novo export..."):
                _progress(bar, status, 10, "Recarregando DETALHADO_CLASSIFICADO do último export")
                df2 = pd.read_excel(last_export, sheet_name="DETALHADO_CLASSIFICADO", engine="openpyxl")

                _progress(bar, status, 25, "Garantindo chaves normalizadas (NOME_*_N)")
                if "NOME_CONTA_N" not in df2.columns or "NOME_PESSOA_N" not in df2.columns or "NOME_CC_N" not in df2.columns:
                    df2 = normalize_keys(df2)

                _progress(bar, status, 35, "Garantindo COMPETENCIA_MES (para consolidar)")
                if "COMPETENCIA_MES" not in df2.columns:
                    if "DATA_EMISSAO" in df2.columns:
                        df2 = add_competencia(df2, competencia_field="DATA_EMISSAO")
                    else:
                        bar.progress(100, text="100% — Falha ❌ (COMPETENCIA_MES/DATA_EMISSAO ausente)")
                        st.error("Não encontrei COMPETENCIA_MES nem DATA_EMISSAO no detalhado para consolidar a DINÂMICA.")
                        st.stop()

                _progress(bar, status, 55, "Reaplicando classificação (histórico SQLite atualizado)")
                df2 = apply_classification(df2)

                _progress(bar, status, 70, "Consolidando DINÂMICA")
                df_con = consolidate_dinamica(df2, value_field="VALOR_TITULO")

                _progress(bar, status, 85, "Exportando Excel (DINÂMICA + Detalhado)")
                out_path = export_excel(df_con, df2, out_dir="data/exports")

                _progress(bar, status, 95, "Aplicando formatação numérica no Excel")
                try:
                    apply_excel_number_format(out_path)
                except Exception:
                    pass

                elapsed = time.perf_counter() - t0
                bar.progress(100, text=f"100% — Concluído ✅ (tempo: {elapsed:.1f}s)")
                status.success("Reexportação concluída com sucesso.")

                st.session_state["ultimo_export_reclassificado"] = out_path
                st.success(f"✅ Novo export gerado: {out_path.name}")

                df_con_view = df_con.copy()
                if "VALOR" in df_con_view.columns:
                    df_con_view["VALOR_BR"] = df_con_view["VALOR"].apply(fmt_br)

                st.subheader("Prévia DINÂMICA (valores BR - visual)")
                cols = [c for c in ["COMPETENCIA_MES", "CLASSIFICACAO_RF", "VALOR_BR"] if c in df_con_view.columns]
                if cols:
                    st.dataframe(df_con_view[cols], use_container_width=True)

        except Exception as e:
            bar.progress(100, text="100% — Falha ❌ (veja o erro abaixo)")
            status.error("Ocorreu um erro durante a reexportação.")
            st.exception(e)

    if st.session_state.get("ultimo_export_reclassificado"):
        out_path = st.session_state["ultimo_export_reclassificado"]
        try:
            with open(out_path, "rb") as f:
                st.download_button(
                    "⬇️ Baixar último Excel reexportado",
                    data=f,
                    file_name=Path(out_path).name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception:
            st.warning("Não foi possível abrir o arquivo exportado para download. Reexporte novamente.")

with col2:
    st.write("Dica: comece pelas maiores por VALOR (já está ordenado).")